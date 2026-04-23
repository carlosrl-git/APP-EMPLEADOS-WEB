from datetime import datetime, timedelta
from email.mime.text import MIMEText
from html import escape
from io import BytesIO
import os
import random
import smtplib

from fastapi import FastAPI, Form, Request
from app.ai.routes import router as ai_router
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from passlib.hash import pbkdf2_sha256
from sqlalchemy import create_engine, text
from starlette.middleware.sessions import SessionMiddleware

from slowapi import Limiter
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from slowapi.middleware import SlowAPIMiddleware

from app.core.config import settings


APP_ENV = settings.APP_ENV
SESSION_SECRET = settings.SESSION_SECRET
SESSION_HTTPS_ONLY = settings.SESSION_HTTPS_ONLY

app = FastAPI(title="App Empleados Web")

# Rate limit global: 100 peticiones por minuto por IP
limiter = Limiter(key_func=get_remote_address, default_limits=["100/minute"])
app.state.limiter = limiter


@app.exception_handler(RateLimitExceeded)
async def rate_limit_handler(request: Request, exc: RateLimitExceeded):
    accept = (request.headers.get("accept") or "").lower()

    if "text/html" in accept:
        return HTMLResponse(
            content="""
            <!DOCTYPE html>
            <html lang="es">
            <head>
                <meta charset="UTF-8">
                <title>Demasiadas peticiones</title>
                <style>
                    body{
                        margin:0;
                        background:#0b0b0b;
                        color:#fff;
                        font-family:Arial,sans-serif;
                        display:flex;
                        align-items:center;
                        justify-content:center;
                        min-height:100vh;
                    }
                    .box{
                        max-width:520px;
                        width:90%;
                        background:#151515;
                        border:1px solid #2f2f2f;
                        border-radius:18px;
                        padding:28px;
                        box-shadow:0 20px 50px rgba(0,0,0,.35);
                        text-align:center;
                    }
                    h1{
                        margin:0 0 12px 0;
                        color:#ff4d6d;
                    }
                    p{
                        margin:0;
                        color:#d0d0d0;
                        line-height:1.5;
                    }
                </style>
            </head>
            <body>
                <div class="box">
                    <h1>Demasiadas peticiones</h1>
                    <p>Has superado el límite permitido temporalmente. Espera un momento e inténtalo otra vez.</p>
                </div>
            </body>
            </html>
            """,
            status_code=429,
        )

    return JSONResponse(
        status_code=429,
        content={"detail": "Demasiadas peticiones. Inténtalo más tarde."}
    )


app.add_middleware(SlowAPIMiddleware)
app.include_router(ai_router)

app.add_middleware(
    SessionMiddleware,
    secret_key=SESSION_SECRET,
    same_site="lax",
    https_only=SESSION_HTTPS_ONLY,
    max_age=60 * 60 * 8,
)

app.mount("/static", StaticFiles(directory=os.path.join(os.path.dirname(__file__), "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(os.path.dirname(__file__), "templates"))
engine = create_engine(settings.DATABASE_URL, pool_pre_ping=True, future=True)


@app.middleware("http")
async def security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Cache-Control"] = "no-store"
    response.headers["Pragma"] = "no-cache"
    response.headers["Content-Security-Policy"] = "default-src 'self' https: data: 'unsafe-inline' 'unsafe-eval'; frame-ancestors 'none'; base-uri 'self'; form-action 'self'"
    return response


def render_template(request: Request, name: str, context: dict, status_code: int = 200):
    context["request"] = request
    return templates.TemplateResponse(
    request=request,
    name=name,
    context=context,
    status_code=status_code,
   )

def send_reset_email(to_email: str, code: str) -> bool:
    smtp_host = os.getenv("SMTP_HOST")
    smtp_port = os.getenv("SMTP_PORT")
    smtp_user = os.getenv("SMTP_USER")
    smtp_password = os.getenv("SMTP_PASSWORD")
    from_email = os.getenv("SMTP_FROM") or smtp_user

    if not all([smtp_host, smtp_port, smtp_user, smtp_password, from_email]):
        print("[CORREO] Variables SMTP incompletas. No se envió el correo.")
        return False

    msg = MIMEText(
        f"Tu código de recuperación es: {code}\n\n"
        "Si no solicitaste este cambio, puedes ignorar este correo.",
        "plain",
        "utf-8",
    )
    msg["Subject"] = "Recuperación de contraseña - AppEmpleados"
    msg["From"] = from_email
    msg["To"] = to_email

    try:
        server = smtplib.SMTP(smtp_host, int(smtp_port), timeout=20)
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(smtp_user, smtp_password)
        server.sendmail(from_email, [to_email], msg.as_string())
        server.quit()
        print(f"[CORREO] Email enviado correctamente a {to_email}")
        return True
    except Exception as e:
        print(f"[CORREO ERROR] {e}")
        return False


def send_ticket_email(ticket: dict) -> bool:
    smtp_host = os.getenv("SMTP_HOST")
    smtp_port = os.getenv("SMTP_PORT")
    smtp_user = os.getenv("SMTP_USER")
    smtp_password = os.getenv("SMTP_PASSWORD")
    from_email = os.getenv("SMTP_FROM") or smtp_user
    support_to = (
        os.getenv("TICKETS_TO_EMAIL")
        or os.getenv("SUPPORT_EMAIL")
        or os.getenv("PRESUPUESTO_EMAIL")
        or from_email
    )

    if not all([smtp_host, smtp_port, smtp_user, smtp_password, from_email, support_to]):
        print("[TICKET CORREO] Variables SMTP incompletas. No se envió el aviso.")
        return False

    body = (
        "Se ha creado un nuevo ticket desde la app.\n\n"
        f"ID: {ticket.get('id')}\n"
        f"Usuario: {ticket.get('nombre_usuario')}\n"
        f"Email: {ticket.get('email_usuario')}\n"
        f"Título: {ticket.get('titulo')}\n"
        f"Descripción: {ticket.get('descripcion')}\n"
        f"Tipo: {ticket.get('tipo')}\n"
        f"Prioridad: {ticket.get('prioridad')}\n"
        f"Estado: {ticket.get('estado')}\n"
        f"Fecha: {ticket.get('fecha_creacion')}\n"
    )

    msg = MIMEText(body, "plain", "utf-8")
    msg["Subject"] = f"[Ticket #{ticket.get('id')}] {ticket.get('titulo')}"
    msg["From"] = from_email
    msg["To"] = support_to
    if ticket.get("email_usuario"):
        msg["Reply-To"] = ticket["email_usuario"]

    try:
        server = smtplib.SMTP(smtp_host, int(smtp_port), timeout=20)
        server.ehlo()
        server.starttls()
        server.ehlo()
        server.login(smtp_user, smtp_password)
        server.sendmail(from_email, [support_to], msg.as_string())
        server.quit()
        print(f"[TICKET CORREO] Ticket enviado correctamente a {support_to}")
        return True
    except Exception as e:
        print(f"[TICKET CORREO ERROR] {e}")
        return False


def get_client_ip(request: Request) -> str:
    forwarded = request.headers.get("x-forwarded-for")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.client.host if request.client else "unknown"


def get_current_username(request: Request):
    username = request.session.get("username")
    if username:
        return username

    username = request.session.get("session_user")
    if username:
        return username

    legacy_cookie = request.cookies.get("session_user")
    return legacy_cookie


def get_current_role(username: str):
    if not username:
        return None
    with engine.begin() as conn:
        user = conn.execute(
            text("SELECT rol FROM usuarios WHERE username = :u AND activo = true"),
            {"u": username}
        ).mappings().first()
    return user["rol"] if user else None


def get_current_empresa_id(request: Request):
    empresa_id = request.session.get("empresa_id")
    if empresa_id is None:
        return None
    return int(empresa_id)


def require_login(request: Request):
    username = get_current_username(request)
    if not username:
        return None, None, RedirectResponse(url="/", status_code=303)

    rol = request.session.get("rol") or request.session.get("session_role")
    if not rol:
        rol = get_current_role(username)

    if not rol:
        request.session.clear()
        response = RedirectResponse(url="/", status_code=303)
        response.delete_cookie("session_user")
        return None, None, response

    return username, rol, None


def require_admin(request: Request):
    username, rol, response = require_login(request)
    if response:
        return None, None, response
    if rol != "admin":
        return username, rol, RedirectResponse(url="/dashboard", status_code=303)
    return username, rol, None


@app.get("/", response_class=HTMLResponse)
def login_page(request: Request):
    username = get_current_username(request)
    if username and get_current_role(username):
        return RedirectResponse(url="/dashboard", status_code=303)
    request.session.clear()
    response = render_template(request, "login.html", {"error": ""})
    response.delete_cookie("session_user")
    return response


@app.get("/login")
def login_redirect():
    return RedirectResponse(url="/", status_code=303)


@app.post("/login", response_class=HTMLResponse)
@limiter.limit("10/minute")
def login(request: Request, username: str = Form(...), password: str = Form(...)):
    now = datetime.utcnow()
    username = (username or "").strip()
    password = password or ""
    ip = get_client_ip(request)

    if not username or not password:
        return render_template(request, "login.html", {"error": "Usuario o contraseña incorrectos"}, 401)

    try:
        with engine.begin() as conn:
            attempt = conn.execute(
                text("""
                    SELECT username, attempts, last_attempt
                    FROM login_attempts
                    WHERE username = :u
                """),
                {"u": username},
            ).mappings().first()

            if (
                attempt
                and attempt["attempts"] >= 5
                and attempt["last_attempt"] is not None
                and (now - attempt["last_attempt"]).total_seconds() < 300
            ):
                return render_template(
                    request,
                    "login.html",
                    {"error": "Usuario bloqueado 5 minutos por demasiados intentos"},
                    429,
                )

            user = conn.execute(
                text("""
                    SELECT id, username, email, password_hash, rol, activo, empresa_id
                    FROM usuarios
                    WHERE username = :u
                """),
                {"u": username},
            ).mappings().first()

            if user and user["activo"] and pbkdf2_sha256.verify(password, user["password_hash"]):
                if user["empresa_id"] is None:
                    return render_template(
                        request,
                        "login.html",
                        {"error": "Usuario sin empresa asignada. Contacta con el administrador."},
                        403,
                    )

                conn.execute(
                    text("DELETE FROM login_attempts WHERE username = :u"),
                    {"u": username},
                )

                request.session.clear()

                # Compatibilidad con el sistema antiguo
                request.session["session_user"] = user["username"]
                request.session["session_role"] = user["rol"]
                request.session["client_ip"] = ip

                # Nuevo sistema para multiempresa
                request.session["user_id"] = user["id"]
                request.session["username"] = user["username"]
                request.session["rol"] = user["rol"]
                request.session["empresa_id"] = user["empresa_id"]
                request.session["nivel"] = 100 if user["rol"] == "admin" else 70

                response = RedirectResponse(url="/dashboard", status_code=303)
                response.delete_cookie("session_user")
                return response

            if attempt:
                conn.execute(
                    text("UPDATE login_attempts SET attempts = attempts + 1, last_attempt = :t WHERE username = :u"),
                    {"t": now, "u": username},
                )
            else:
                conn.execute(
                    text("INSERT INTO login_attempts (username, attempts, last_attempt) VALUES (:u, 1, :t)"),
                    {"u": username, "t": now},
                )

        return render_template(request, "login.html", {"error": "Usuario o contraseña incorrectos"}, 401)

    except Exception as e:
        print(f"[LOGIN ERROR] {e}")
        return render_template(request, "login.html", {"error": "Error interno al iniciar sesión"}, 500)


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request):
    username, rol, response = require_login(request)
    if response:
        return response

    with engine.begin() as conn:
        total_users = conn.execute(text("SELECT COUNT(*) FROM usuarios")).scalar()
        blocked = conn.execute(text("SELECT COUNT(*) FROM login_attempts WHERE attempts >= 5")).scalar()
        resets = conn.execute(text("SELECT COUNT(*) FROM password_resets WHERE used = false")).scalar()
        total_trabajadores = conn.execute(text("SELECT COUNT(*) FROM trabajadores")).scalar()
        incidencias_abiertas = conn.execute(text("SELECT COUNT(*) FROM incidencias WHERE estado = 'abierta'")).scalar()
        tareas_pendientes = conn.execute(
            text("SELECT COUNT(*) FROM tareas WHERE estado IN ('pendiente','en_progreso')")
        ).scalar()

    return render_template(request, "dashboard.html", {
        "username": username,
        "rol": rol,
        "total_users": total_users,
        "blocked": blocked,
        "resets": resets,
        "total_trabajadores": total_trabajadores,
        "incidencias_abiertas": incidencias_abiertas,
        "tareas_pendientes": tareas_pendientes,
        "active_page": "inicio",
    })


@app.get("/trabajadores", response_class=HTMLResponse)
def ver_trabajadores(request: Request, q: str = "", departamento_id: str = ""):
    username, rol, response = require_login(request)
    if response:
        return response

    q = (q or "").strip()
    departamento_id = (departamento_id or "").strip()

    sql = """
        SELECT t.id, t.nombre, t.apellidos, t.puesto, t.email, d.nombre AS departamento
        FROM trabajadores t
        LEFT JOIN departamentos d ON d.id = t.departamento_id
        WHERE 1=1
    """
    params = {}

    if q:
        sql += " AND (LOWER(t.nombre) LIKE LOWER(:q) OR LOWER(COALESCE(t.apellidos, '')) LIKE LOWER(:q))"
        params["q"] = f"%{q}%"

    if departamento_id:
        try:
            params["departamento_id"] = int(departamento_id)
            sql += " AND t.departamento_id = :departamento_id"
        except ValueError:
            departamento_id = ""

    sql += " ORDER BY t.id"

    with engine.begin() as conn:
        trabajadores = conn.execute(text(sql), params).mappings().all()
        departamentos = conn.execute(
            text("SELECT id, nombre FROM departamentos WHERE activo = true ORDER BY nombre")
        ).mappings().all()

    return render_template(request, "trabajadores.html", {
        "username": username,
        "rol": rol,
        "trabajadores": trabajadores,
        "departamentos": departamentos,
        "q": q,
        "departamento_id": departamento_id,
        "active_page": "trabajadores",
    })


@app.get("/trabajadores/nuevo", response_class=HTMLResponse)
def nuevo_trabajador_page(request: Request):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        departamentos = conn.execute(
            text("SELECT id, nombre FROM departamentos WHERE activo = true ORDER BY nombre")
        ).mappings().all()

    return render_template(request, "trabajador_nuevo.html", {
        "username": username,
        "rol": rol,
        "departamentos": departamentos,
        "error": "",
        "msg": "",
        "active_page": "trabajadores",
    })


@app.post("/trabajadores/guardar2")
def guardar_trabajador_2(
    request: Request,
    nombre: str = Form(...),
    apellidos: str = Form(""),
    email: str = Form(""),
    telefono: str = Form(""),
    puesto: str = Form(""),
    departamento_id: str = Form(""),
):
    username, rol, response = require_admin(request)
    if response:
        return response

    try:
        dep_id = int(departamento_id) if str(departamento_id).strip() else None

        with engine.begin() as conn:
            conn.execute(
                text("""
                    INSERT INTO trabajadores
                    (nombre, apellidos, email, telefono, puesto, departamento_id, activo)
                    VALUES (:nombre, :apellidos, :email, :telefono, :puesto, :departamento_id, true)
                """),
                {
                    "nombre": nombre.strip(),
                    "apellidos": apellidos.strip() or None,
                    "email": email.strip() or None,
                    "telefono": telefono.strip() or None,
                    "puesto": puesto.strip() or None,
                    "departamento_id": dep_id,
                },
            )

        return RedirectResponse(url="/trabajadores", status_code=303)

    except Exception as e:
        print(f"[GUARDAR TRABAJADOR 2 ERROR] {e}")
        return RedirectResponse(url="/trabajadores/nuevo", status_code=303)


@app.get("/trabajadores/{id}", response_class=HTMLResponse)
def ficha_trabajador(request: Request, id: int):
    username, rol, response = require_login(request)
    if response:
        return response

    with engine.begin() as conn:
        t = conn.execute(text("""
            SELECT t.*, d.nombre AS departamento
            FROM trabajadores t
            LEFT JOIN departamentos d ON d.id = t.departamento_id
            WHERE t.id = :id
        """), {"id": id}).mappings().first()

        if not t:
            return RedirectResponse(url="/trabajadores", status_code=303)

        ausencias = conn.execute(text("""
            SELECT id, tipo, fecha_inicio, fecha_fin, motivo, estado
            FROM ausencias
            WHERE trabajador_id = :id
            ORDER BY fecha_inicio DESC
        """), {"id": id}).mappings().all()

        horas = conn.execute(text("""
            SELECT id, fecha, horas, horas_extra, tipo
            FROM horas_trabajadas
            WHERE trabajador_id = :id
            ORDER BY fecha DESC
        """), {"id": id}).mappings().all()

        incidencias = conn.execute(text("""
            SELECT id, descripcion, estado, fecha
            FROM incidencias
            WHERE trabajador_id = :id
            ORDER BY fecha DESC
        """), {"id": id}).mappings().all()

        tareas = conn.execute(text("""
            SELECT id, descripcion, estado, resultado, evaluacion, fecha_asignacion, fecha_realizacion
            FROM tareas_extra
            WHERE trabajador_id = :id
            ORDER BY fecha_asignacion DESC
        """), {"id": id}).mappings().all()

        turnos = conn.execute(text("""
            SELECT tu.id, tu.fecha, tu.turno, tu.hora_inicio, tu.hora_fin, tu.estado, tu.observaciones, z.nombre AS zona
            FROM turnos tu
            LEFT JOIN zonas z ON tu.zona_id = z.id
            WHERE tu.trabajador_id = :id
            ORDER BY tu.fecha DESC, tu.hora_inicio DESC
        """), {"id": id}).mappings().all()

        rutas = conn.execute(text("""
            SELECT r.id, r.fecha, r.hora_inicio_jornada, r.hora_fin_jornada, r.total_horas, r.observaciones_generales
            FROM rutas r
            WHERE r.trabajador_id = :id
            ORDER BY r.fecha DESC, r.hora_inicio_jornada ASC
        """), {"id": id}).mappings().all()

    return render_template(request, "trabajador_ficha.html", {
        "username": username,
        "rol": rol,
        "t": t,
        "ausencias": ausencias,
        "horas": horas,
        "incidencias": incidencias,
        "tareas": tareas,
        "turnos": turnos,
        "rutas": rutas,
        "active_page": "trabajadores"
    })


@app.get("/trabajadores/editar/{id}", response_class=HTMLResponse)
def editar_trabajador_page(request: Request, id: int):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        t = conn.execute(text("SELECT * FROM trabajadores WHERE id = :id"), {"id": id}).mappings().first()
        departamentos = conn.execute(
            text("SELECT id, nombre FROM departamentos WHERE activo = true")
        ).mappings().all()

    return render_template(request, "trabajador_editar.html", {
        "username": username,
        "rol": rol,
        "t": t,
        "departamentos": departamentos,
        "active_page": "trabajadores",
    })


@app.post("/trabajadores/editar/{id}")
def editar_trabajador(
    request: Request,
    id: int,
    nombre: str = Form(...),
    apellidos: str = Form(""),
    email: str = Form(""),
    telefono: str = Form(""),
    puesto: str = Form(""),
    departamento_id: str = Form(""),
):
    username, rol, response = require_admin(request)
    if response:
        return response

    dep_id = int(departamento_id) if str(departamento_id).strip() else None

    with engine.begin() as conn:
        conn.execute(
            text("""
                UPDATE trabajadores SET
                    nombre=:nombre,
                    apellidos=:apellidos,
                    email=:email,
                    telefono=:telefono,
                    puesto=:puesto,
                    departamento_id=:dep
                WHERE id=:id
            """),
            {
                "id": id,
                "nombre": nombre.strip(),
                "apellidos": apellidos.strip() or None,
                "email": email.strip() or None,
                "telefono": telefono.strip() or None,
                "puesto": puesto.strip() or None,
                "dep": dep_id,
            },
        )

    return RedirectResponse(url="/trabajadores", status_code=303)


@app.get("/trabajadores/eliminar/{id}")
def eliminar_trabajador(request: Request, id: int):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        conn.execute(text("DELETE FROM trabajadores WHERE id = :id"), {"id": id})

    return RedirectResponse(url="/trabajadores", status_code=303)


@app.get("/incidencias", response_class=HTMLResponse)
def ver_incidencias(request: Request, estado: str = ""):
    username, rol, response = require_login(request)
    if response:
        return response

    estado = (estado or "").strip()

    sql = """
        SELECT i.id, i.descripcion, i.estado, i.fecha, t.nombre, t.apellidos
        FROM incidencias i
        LEFT JOIN trabajadores t ON t.id = i.trabajador_id
        WHERE 1=1
    """
    params = {}

    if estado:
        sql += " AND i.estado = :estado"
        params["estado"] = estado

    sql += " ORDER BY i.fecha DESC"

    with engine.begin() as conn:
        incidencias = conn.execute(text(sql), params).mappings().all()

    return render_template(request, "incidencias.html", {
        "username": username,
        "rol": rol,
        "incidencias": incidencias,
        "estado": estado,
        "active_page": "incidencias",
    })


@app.get("/incidencias/nueva", response_class=HTMLResponse)
def nueva_incidencia_page(request: Request):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        trabajadores = conn.execute(
            text("SELECT id, nombre, apellidos FROM trabajadores ORDER BY nombre, apellidos")
        ).mappings().all()

    return render_template(request, "incidencia_nueva.html", {
        "username": username,
        "rol": rol,
        "trabajadores": trabajadores,
        "error": "",
        "msg": "",
        "active_page": "incidencias",
    })


@app.post("/incidencias/nueva")
def guardar_incidencia(
    request: Request,
    descripcion: str = Form(...),
    trabajador_id: str = Form(""),
    estado: str = Form("abierta"),
):
    username, rol, response = require_admin(request)
    if response:
        return response

    try:
        trabajador = int(trabajador_id) if str(trabajador_id).strip() else None

        with engine.begin() as conn:
            conn.execute(
                text("INSERT INTO incidencias (descripcion, trabajador_id, estado) VALUES (:descripcion, :trabajador_id, :estado)"),
                {
                    "descripcion": descripcion.strip(),
                    "trabajador_id": trabajador,
                    "estado": estado.strip() or "abierta",
                },
            )

        return RedirectResponse(url="/incidencias", status_code=303)

    except Exception as e:
        print(f"[NUEVA INCIDENCIA ERROR] {e}")
        with engine.begin() as conn:
            trabajadores = conn.execute(
                text("SELECT id, nombre, apellidos FROM trabajadores ORDER BY nombre, apellidos")
            ).mappings().all()
        return render_template(request, "incidencia_nueva.html", {
            "username": username,
            "rol": rol,
            "trabajadores": trabajadores,
            "error": "Error al crear la incidencia",
            "msg": "",
            "active_page": "incidencias",
        })


@app.get("/incidencias/estado/{id}/{nuevo_estado}")
def cambiar_estado_incidencia(request: Request, id: int, nuevo_estado: str):
    username, rol, response = require_admin(request)
    if response:
        return response

    estados_validos = ["abierta", "en_proceso", "completada"]

    if nuevo_estado not in estados_validos:
        return RedirectResponse(url="/incidencias", status_code=303)

    with engine.begin() as conn:
        conn.execute(
            text("UPDATE incidencias SET estado = :estado WHERE id = :id"),
            {"estado": nuevo_estado, "id": id},
        )

    return RedirectResponse(url="/incidencias", status_code=303)


@app.get("/tickets", response_class=HTMLResponse)
def ver_tickets(request: Request, ok: str = "", error: str = ""):
    username, rol, response = require_login(request)
    if response:
        return response

    with engine.begin() as conn:
        usuario_actual = conn.execute(
            text("SELECT username, email, rol FROM usuarios WHERE username = :u"),
            {"u": username},
        ).mappings().first()

        email_actual = (usuario_actual["email"] or "").strip() if usuario_actual else ""

        if rol == "admin":
            tickets = conn.execute(
                text("""
                    SELECT id, nombre_usuario, email_usuario, titulo, descripcion, tipo, prioridad, estado, fecha_creacion
                    FROM tickets
                    ORDER BY id DESC
                """)
            ).mappings().all()
        else:
            tickets = conn.execute(
                text("""
                    SELECT id, nombre_usuario, email_usuario, titulo, descripcion, tipo, prioridad, estado, fecha_creacion
                    FROM tickets
                    WHERE LOWER(email_usuario) = LOWER(:email_usuario)
                    ORDER BY id DESC
                """),
                {"email_usuario": email_actual},
            ).mappings().all()

    ok_html = ""
    error_html = ""

    if ok:
        ok_html = '<div style="margin-bottom:16px;padding:12px 16px;background:#133a1b;border:1px solid #1e6b2d;border-radius:10px;color:#d8ffd8;">Ticket creado correctamente.</div>'
    if error:
        error_html = f'<div style="margin-bottom:16px;padding:12px 16px;background:#3a1313;border:1px solid #6b1e1e;border-radius:10px;color:#ffd8d8;">{escape(error)}</div>'

    filas = ""
    for t in tickets:
        filas += f"""
        <tr>
            <td style="padding:10px;border-bottom:1px solid #2a2a2a;">{t['id']}</td>
            <td style="padding:10px;border-bottom:1px solid #2a2a2a;">{escape(t['nombre_usuario'] or '')}</td>
            <td style="padding:10px;border-bottom:1px solid #2a2a2a;">{escape(t['email_usuario'] or '')}</td>
            <td style="padding:10px;border-bottom:1px solid #2a2a2a;">{escape(t['titulo'] or '')}</td>
            <td style="padding:10px;border-bottom:1px solid #2a2a2a;">{escape((t['descripcion'] or '')[:120])}</td>
            <td style="padding:10px;border-bottom:1px solid #2a2a2a;">{escape(t['tipo'] or '')}</td>
            <td style="padding:10px;border-bottom:1px solid #2a2a2a;">{escape(t['prioridad'] or '')}</td>
            <td style="padding:10px;border-bottom:1px solid #2a2a2a;">{escape(t['estado'] or '')}</td>
            <td style="padding:10px;border-bottom:1px solid #2a2a2a;">{escape(str(t['fecha_creacion'] or ''))}</td>
        </tr>
        """

    if not filas:
        filas = """
        <tr>
            <td colspan="9" style="padding:16px;border-bottom:1px solid #2a2a2a;">No hay tickets registrados.</td>
        </tr>
        """

    html = f"""
    <!DOCTYPE html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <title>Tickets | App Empleados</title>
        <link rel="stylesheet" href="/static/styles.css">
    </head>
    <body style="background:#0f0f0f;color:#fff;font-family:Arial,sans-serif;margin:0;padding:0;">
        <div style="max-width:1400px;margin:0 auto;padding:24px;">
            <div style="display:flex;justify-content:space-between;align-items:center;gap:12px;flex-wrap:wrap;margin-bottom:24px;">
                <div>
                    <h1 style="margin:0 0 6px 0;">Soporte / Tickets</h1>
                    <p style="margin:0;color:#bdbdbd;">Crea incidencias o peticiones sobre la aplicación.</p>
                </div>
                <div style="display:flex;gap:10px;flex-wrap:wrap;">
                    <a href="/dashboard" style="text-decoration:none;padding:10px 16px;background:#1f1f1f;color:#fff;border:1px solid #333;border-radius:10px;">Volver</a>
                    <a href="/logout" style="text-decoration:none;padding:10px 16px;background:#1f1f1f;color:#fff;border:1px solid #333;border-radius:10px;">Salir</a>
                </div>
            </div>

            {ok_html}
            {error_html}

            <div style="background:#171717;border:1px solid #2a2a2a;border-radius:16px;padding:20px;margin-bottom:24px;">
                <h2 style="margin-top:0;">Nuevo ticket</h2>
                <form method="post" action="/tickets/nuevo" style="display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:14px;">
                    <div>
                        <label style="display:block;margin-bottom:6px;">Nombre</label>
                        <input name="nombre_usuario" value="{escape(username)}" required
                               style="width:100%;padding:12px;border-radius:10px;border:1px solid #333;background:#101010;color:#fff;">
                    </div>
                    <div>
                        <label style="display:block;margin-bottom:6px;">Email</label>
                        <input name="email_usuario" value="{escape(email_actual)}" required
                               style="width:100%;padding:12px;border-radius:10px;border:1px solid #333;background:#101010;color:#fff;">
                    </div>
                    <div>
                        <label style="display:block;margin-bottom:6px;">Tipo</label>
                        <select name="tipo"
                                style="width:100%;padding:12px;border-radius:10px;border:1px solid #333;background:#101010;color:#fff;">
                            <option value="peticion">Petición</option>
                            <option value="incidencia">Incidencia</option>
                            <option value="mejora">Mejora</option>
                            <option value="error">Error</option>
                            <option value="soporte">Soporte</option>
                        </select>
                    </div>
                    <div>
                        <label style="display:block;margin-bottom:6px;">Prioridad</label>
                        <select name="prioridad"
                                style="width:100%;padding:12px;border-radius:10px;border:1px solid #333;background:#101010;color:#fff;">
                            <option value="baja">Baja</option>
                            <option value="media" selected>Media</option>
                            <option value="alta">Alta</option>
                            <option value="urgente">Urgente</option>
                        </select>
                    </div>
                    <div style="grid-column:1/-1;">
                        <label style="display:block;margin-bottom:6px;">Título</label>
                        <input name="titulo" required
                               style="width:100%;padding:12px;border-radius:10px;border:1px solid #333;background:#101010;color:#fff;">
                    </div>
                    <div style="grid-column:1/-1;">
                        <label style="display:block;margin-bottom:6px;">Descripción</label>
                        <textarea name="descripcion" rows="6" required
                                  style="width:100%;padding:12px;border-radius:10px;border:1px solid #333;background:#101010;color:#fff;resize:vertical;"></textarea>
                    </div>
                    <div style="grid-column:1/-1;">
                        <button type="submit"
                                style="padding:12px 18px;background:#c00000;color:#fff;border:none;border-radius:10px;cursor:pointer;font-weight:700;">
                            Crear ticket
                        </button>
                    </div>
                </form>
            </div>

            <div style="background:#171717;border:1px solid #2a2a2a;border-radius:16px;padding:20px;">
                <h2 style="margin-top:0;">Listado de tickets</h2>
                <div style="overflow:auto;">
                    <table style="width:100%;border-collapse:collapse;min-width:1100px;">
                        <thead>
                            <tr style="background:#101010;">
                                <th style="padding:10px;text-align:left;border-bottom:1px solid #333;">ID</th>
                                <th style="padding:10px;text-align:left;border-bottom:1px solid #333;">Usuario</th>
                                <th style="padding:10px;text-align:left;border-bottom:1px solid #333;">Email</th>
                                <th style="padding:10px;text-align:left;border-bottom:1px solid #333;">Título</th>
                                <th style="padding:10px;text-align:left;border-bottom:1px solid #333;">Descripción</th>
                                <th style="padding:10px;text-align:left;border-bottom:1px solid #333;">Tipo</th>
                                <th style="padding:10px;text-align:left;border-bottom:1px solid #333;">Prioridad</th>
                                <th style="padding:10px;text-align:left;border-bottom:1px solid #333;">Estado</th>
                                <th style="padding:10px;text-align:left;border-bottom:1px solid #333;">Fecha</th>
                            </tr>
                        </thead>
                        <tbody>
                            {filas}
                        </tbody>
                    </table>
                </div>
            </div>
        </div>
    </body>
    </html>
    """

    return HTMLResponse(content=html)


@app.post("/tickets/nuevo")
@limiter.limit("15/minute")
def crear_ticket_web(
    request: Request,
    nombre_usuario: str = Form(...),
    email_usuario: str = Form(...),
    titulo: str = Form(...),
    descripcion: str = Form(...),
    tipo: str = Form("peticion"),
    prioridad: str = Form("media"),
):
    username, rol, response = require_login(request)
    if response:
        return response

    nombre_usuario = (nombre_usuario or "").strip()
    email_usuario = (email_usuario or "").strip()
    titulo = (titulo or "").strip()
    descripcion = (descripcion or "").strip()
    tipo = (tipo or "peticion").strip().lower()
    prioridad = (prioridad or "media").strip().lower()

    tipos_validos = {"peticion", "incidencia", "mejora", "error", "soporte"}
    prioridades_validas = {"baja", "media", "alta", "urgente"}

    if not nombre_usuario or not email_usuario or not titulo or not descripcion:
        return RedirectResponse(url="/tickets?error=Todos+los+campos+son+obligatorios", status_code=303)

    if tipo not in tipos_validos:
        tipo = "peticion"

    if prioridad not in prioridades_validas:
        prioridad = "media"

    try:
        with engine.begin() as conn:
            ticket = conn.execute(
                text("""
                    INSERT INTO tickets (
                        nombre_usuario,
                        email_usuario,
                        titulo,
                        descripcion,
                        tipo,
                        prioridad
                    )
                    VALUES (
                        :nombre_usuario,
                        :email_usuario,
                        :titulo,
                        :descripcion,
                        :tipo,
                        :prioridad
                    )
                    RETURNING id, nombre_usuario, email_usuario, titulo, descripcion, tipo, prioridad, estado, fecha_creacion
                """),
                {
                    "nombre_usuario": nombre_usuario,
                    "email_usuario": email_usuario,
                    "titulo": titulo,
                    "descripcion": descripcion,
                    "tipo": tipo,
                    "prioridad": prioridad,
                }
            ).mappings().first()

        if ticket:
            send_ticket_email(dict(ticket))

        return RedirectResponse(url="/tickets?ok=1", status_code=303)

    except Exception as e:
        print(f"[TICKET NUEVO ERROR] {e}")
        return RedirectResponse(url="/tickets?error=Error+al+crear+el+ticket", status_code=303)


@app.get("/tareas", response_class=HTMLResponse)
def ver_tareas(request: Request, estado: str = "", prioridad: str = ""):
    username, rol, response = require_login(request)
    if response:
        return response

    estado = (estado or "").strip()
    prioridad = (prioridad or "").strip()

    sql = """
        SELECT ta.id, ta.titulo, ta.estado, ta.prioridad, ta.fecha_asignacion, ta.fecha_vencimiento,
               tr.nombre, tr.apellidos
        FROM tareas ta
        LEFT JOIN trabajadores tr ON tr.id = ta.trabajador_id
        WHERE 1=1
    """
    params = {}

    if estado:
        sql += " AND ta.estado = :estado"
        params["estado"] = estado

    if prioridad:
        sql += " AND ta.prioridad = :prioridad"
        params["prioridad"] = prioridad

    sql += " ORDER BY ta.fecha_asignacion DESC, ta.id DESC"

    with engine.begin() as conn:
        tareas = conn.execute(text(sql), params).mappings().all()

    return render_template(request, "tareas.html", {
        "username": username,
        "rol": rol,
        "tareas": tareas,
        "estado": estado,
        "prioridad": prioridad,
        "active_page": "tareas",
    })


@app.get("/tareas/nueva", response_class=HTMLResponse)
def nueva_tarea_page(request: Request):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        trabajadores = conn.execute(
            text("SELECT id, nombre, apellidos FROM trabajadores ORDER BY nombre, apellidos")
        ).mappings().all()

    return render_template(request, "tarea_nueva.html", {
        "username": username,
        "rol": rol,
        "trabajadores": trabajadores,
        "error": "",
        "msg": "",
        "active_page": "tareas",
    })


@app.post("/tareas/nueva")
def guardar_tarea(
    request: Request,
    titulo: str = Form(...),
    descripcion: str = Form(""),
    trabajador_id: str = Form(...),
    prioridad: str = Form("media"),
    estado: str = Form("pendiente"),
    fecha_asignacion: str = Form(...),
    fecha_vencimiento: str = Form(""),
):
    username, rol, response = require_admin(request)
    if response:
        return response

    try:
        with engine.begin() as conn:
            conn.execute(
                text("""
                    INSERT INTO tareas (
                        trabajador_id, titulo, descripcion, prioridad, estado,
                        fecha_asignacion, fecha_vencimiento
                    ) VALUES (
                        :trabajador_id, :titulo, :descripcion, :prioridad, :estado,
                        :fecha_asignacion, :fecha_vencimiento
                    )
                """),
                {
                    "trabajador_id": int(trabajador_id),
                    "titulo": titulo.strip(),
                    "descripcion": descripcion.strip() or None,
                    "prioridad": prioridad.strip() or "media",
                    "estado": estado.strip() or "pendiente",
                    "fecha_asignacion": fecha_asignacion,
                    "fecha_vencimiento": fecha_vencimiento or None,
                },
            )

        return RedirectResponse(url="/tareas", status_code=303)

    except Exception as e:
        print(f"[NUEVA TAREA ERROR] {e}")
        with engine.begin() as conn:
            trabajadores = conn.execute(
                text("SELECT id, nombre, apellidos FROM trabajadores ORDER BY nombre, apellidos")
            ).mappings().all()
        return render_template(request, "tarea_nueva.html", {
            "username": username,
            "rol": rol,
            "trabajadores": trabajadores,
            "error": "Error al crear la tarea",
            "msg": "",
            "active_page": "tareas",
        })


@app.get("/tareas/estado/{id}/{nuevo_estado}")
def cambiar_estado_tarea(request: Request, id: int, nuevo_estado: str):
    username, rol, response = require_admin(request)
    if response:
        return response

    estados_validos = ["pendiente", "en_progreso", "finalizada", "cancelada"]

    if nuevo_estado not in estados_validos:
        return RedirectResponse(url="/tareas", status_code=303)

    with engine.begin() as conn:
        conn.execute(
            text("UPDATE tareas SET estado = :estado WHERE id = :id"),
            {"estado": nuevo_estado, "id": id},
        )

    return RedirectResponse(url="/tareas", status_code=303)


@app.get("/tareas/editar/{id}", response_class=HTMLResponse)
def editar_tarea_page(request: Request, id: int):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        tarea = conn.execute(text("SELECT * FROM tareas WHERE id = :id"), {"id": id}).mappings().first()
        trabajadores = conn.execute(
            text("SELECT id, nombre, apellidos FROM trabajadores ORDER BY nombre")
        ).mappings().all()

    return render_template(request, "tarea_editar.html", {
        "username": username,
        "rol": rol,
        "t": tarea,
        "trabajadores": trabajadores,
        "active_page": "tareas",
    })


@app.post("/tareas/editar/{id}")
def editar_tarea(
    request: Request,
    id: int,
    titulo: str = Form(...),
    descripcion: str = Form(""),
    trabajador_id: str = Form(...),
    prioridad: str = Form(...),
    estado: str = Form(...),
    fecha_asignacion: str = Form(...),
    fecha_vencimiento: str = Form(""),
):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE tareas SET
                titulo=:titulo,
                descripcion=:descripcion,
                trabajador_id=:trabajador_id,
                prioridad=:prioridad,
                estado=:estado,
                fecha_asignacion=:fecha_asignacion,
                fecha_vencimiento=:fecha_vencimiento
            WHERE id=:id
        """), {
            "id": id,
            "titulo": titulo.strip(),
            "descripcion": descripcion.strip() or None,
            "trabajador_id": int(trabajador_id),
            "prioridad": prioridad.strip(),
            "estado": estado.strip(),
            "fecha_asignacion": fecha_asignacion,
            "fecha_vencimiento": fecha_vencimiento or None,
        })

    return RedirectResponse(url="/tareas", status_code=303)


@app.get("/tareas/eliminar/{id}")
def eliminar_tarea(request: Request, id: int):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        conn.execute(text("DELETE FROM tareas WHERE id = :id"), {"id": id})

    return RedirectResponse(url="/tareas", status_code=303)


@app.get("/usuarios", response_class=HTMLResponse)
def ver_usuarios(request: Request):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        usuarios = conn.execute(
            text("SELECT id, username, email, rol, activo, creado_en FROM usuarios ORDER BY id")
        ).mappings().all()

    return render_template(request, "usuarios.html", {
        "username": username,
        "rol": rol,
        "usuarios": usuarios,
        "active_page": "usuarios",
    })


@app.get("/usuarios/nuevo", response_class=HTMLResponse)
def nuevo_usuario_page(request: Request):
    username, rol, response = require_admin(request)
    if response:
        return response

    return render_template(request, "usuario_nuevo.html", {
        "username": username,
        "rol": rol,
        "error": "",
        "msg": "",
        "active_page": "usuarios",
    })


@app.post("/usuarios/nuevo")
def guardar_usuario(
    request: Request,
    username_nuevo: str = Form(...),
    email: str = Form(...),
    password: str = Form(...),
    rol: str = Form("usuario"),
):
    username, rol_actual, response = require_admin(request)
    if response:
        return response

    try:
        if len(password) < 8:
            return render_template(request, "usuario_nuevo.html", {
                "username": username,
                "rol": rol_actual,
                "error": "La contraseña debe tener al menos 8 caracteres",
                "msg": "",
                "active_page": "usuarios",
            }, 400)

        password_hash = pbkdf2_sha256.hash(password)

        with engine.begin() as conn:
            conn.execute(
                text("""
                    INSERT INTO usuarios (username, email, password_hash, rol, activo, rol_id)
                    VALUES (:username, :email, :password_hash, :rol, true, 1)
                """),
                {
                    "username": username_nuevo.strip(),
                    "email": email.strip(),
                    "password_hash": password_hash,
                    "rol": rol.strip() or "usuario",
                },
            )

        return RedirectResponse(url="/usuarios", status_code=303)

    except Exception as e:
        print(f"[NUEVO USUARIO ERROR] {e}")
        return render_template(request, "usuario_nuevo.html", {
            "username": username,
            "rol": rol_actual,
            "error": "Error al crear el usuario",
            "msg": "",
            "active_page": "usuarios",
        })


@app.get("/usuarios/editar/{id}", response_class=HTMLResponse)
def editar_usuario_page(request: Request, id: int):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        u = conn.execute(text("SELECT * FROM usuarios WHERE id = :id"), {"id": id}).mappings().first()

    return render_template(request, "usuario_editar.html", {
        "username": username,
        "rol": rol,
        "u": u,
        "active_page": "usuarios",
    })


@app.post("/usuarios/editar/{id}")
def editar_usuario(
    request: Request,
    id: int,
    username_nuevo: str = Form(...),
    email: str = Form(...),
    rol: str = Form(...),
):
    username, rol_actual, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE usuarios SET
                username = :username,
                email = :email,
                rol = :rol
            WHERE id = :id
        """), {
            "id": id,
            "username": username_nuevo.strip(),
            "email": email.strip(),
            "rol": rol.strip(),
        })

    return RedirectResponse(url="/usuarios", status_code=303)


@app.get("/usuarios/toggle/{id}")
def toggle_usuario(request: Request, id: int):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE usuarios
            SET activo = NOT activo
            WHERE id = :id
        """), {"id": id})

    return RedirectResponse(url="/usuarios", status_code=303)


@app.get("/usuarios/eliminar/{id}")
def eliminar_usuario(request: Request, id: int):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        usuario_objetivo = conn.execute(
            text("SELECT id, username, rol FROM usuarios WHERE id = :id"),
            {"id": id}
        ).mappings().first()

        if not usuario_objetivo:
            return RedirectResponse(url="/usuarios", status_code=303)

        if usuario_objetivo["username"] == username:
            return RedirectResponse(url="/usuarios", status_code=303)

        admins_restantes = conn.execute(
            text("SELECT COUNT(*) FROM usuarios WHERE rol = 'admin' AND activo = true AND id <> :id"),
            {"id": id}
        ).scalar()

        if usuario_objetivo["rol"] == "admin" and admins_restantes == 0:
            return RedirectResponse(url="/usuarios", status_code=303)

        conn.execute(
            text("DELETE FROM ai_logs WHERE usuario = :usuario"),
            {"usuario": usuario_objetivo["username"]}
        )

        conn.execute(
            text("DELETE FROM usuarios WHERE id = :id"),
            {"id": id}
        )

    return RedirectResponse(url="/usuarios", status_code=303)


@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    response = RedirectResponse(url="/", status_code=303)
    response.delete_cookie("session_user")
    return response


@app.get("/forgot-password", response_class=HTMLResponse)
def forgot_password_page(request: Request):
    return render_template(request, "forgot_password.html", {"msg": "", "error": ""})


@app.post("/forgot-password", response_class=HTMLResponse)
@limiter.limit("5/minute")
def forgot_password(request: Request, email: str = Form(...)):
    code = str(random.randint(100000, 999999))
    expires_at = datetime.utcnow() + timedelta(minutes=10)

    try:
        with engine.begin() as conn:
            user = conn.execute(
                text("SELECT id FROM usuarios WHERE email = :e AND activo = true"),
                {"e": email.strip()},
            ).mappings().first()

            if user:
                conn.execute(
                    text("DELETE FROM password_resets WHERE email = :e AND used = false"),
                    {"e": email.strip()},
                )
                conn.execute(
                    text("""
                        INSERT INTO password_resets (email, code, expires_at, used)
                        VALUES (:e, :c, :x, false)
                    """),
                    {"e": email.strip(), "c": code, "x": expires_at},
                )
                send_reset_email(email.strip(), code)

        return render_template(
            request,
            "forgot_password.html",
            {"msg": "Si el correo existe, se enviará un código.", "error": ""},
        )

    except Exception as e:
        print(f"[FORGOT PASSWORD ERROR] {e}")
        return render_template(
            request,
            "forgot_password.html",
            {"msg": "", "error": "Error al procesar la solicitud"},
            500,
        )


@app.get("/reset-password", response_class=HTMLResponse)
def reset_password_page(request: Request):
    return render_template(request, "reset_password.html", {"msg": "", "error": ""})


@app.post("/reset-password", response_class=HTMLResponse)
@limiter.limit("5/minute")
def reset_password(
    request: Request,
    email: str = Form(...),
    code: str = Form(...),
    new_password: str = Form(...),
):
    if len(new_password) < 8:
        return render_template(
            request,
            "reset_password.html",
            {"msg": "", "error": "La contraseña debe tener al menos 8 caracteres"},
            400,
        )

    try:
        with engine.begin() as conn:
            row = conn.execute(text("""
                SELECT id, email, code, expires_at, used
                FROM password_resets
                WHERE email = :e AND code = :c AND used = false
                ORDER BY id DESC
                LIMIT 1
            """), {"e": email.strip(), "c": code.strip()}).mappings().first()

            if not row:
                return render_template(
                    request,
                    "reset_password.html",
                    {"msg": "", "error": "Código inválido"},
                    400,
                )

            if datetime.utcnow() > row["expires_at"]:
                return render_template(
                    request,
                    "reset_password.html",
                    {"msg": "", "error": "Código caducado"},
                    400,
                )

            new_hash = pbkdf2_sha256.hash(new_password)
            conn.execute(
                text("UPDATE usuarios SET password_hash = :p WHERE email = :e"),
                {"p": new_hash, "e": email.strip()},
            )
            conn.execute(
                text("UPDATE password_resets SET used = true WHERE id = :id"),
                {"id": row["id"]},
            )

        return render_template(
            request,
            "reset_password.html",
            {"msg": "Contraseña cambiada correctamente", "error": ""},
        )

    except Exception as e:
        print(f"[RESET PASSWORD ERROR] {e}")
        return render_template(
            request,
            "reset_password.html",
            {"msg": "", "error": "Error al cambiar la contraseña"},
            500,
        )


@app.get("/turnos", response_class=HTMLResponse)
def ver_turnos(request: Request):
    username, rol, response = require_login(request)
    if response:
        return response

    with engine.begin() as conn:
        turnos = conn.execute(text("""
            SELECT tu.id, tu.fecha, tu.inicio_jornada, tu.fin_jornada, tu.total_horas, tr.nombre, tr.apellidos
            FROM turnos tu
            LEFT JOIN trabajadores tr ON tr.id = tu.trabajador_id
            ORDER BY tu.fecha DESC
        """)).mappings().all()

        trabajadores = conn.execute(
            text("SELECT id, nombre, apellidos FROM trabajadores ORDER BY nombre")
        ).mappings().all()

    return render_template(request, "turnos.html", {
        "turnos": turnos,
        "trabajadores": trabajadores,
        "username": username,
        "rol": rol,
        "active_page": "turnos",
    })


@app.post("/turnos/nuevo")
def crear_turno(
    request: Request,
    trabajador_id: str = Form(...),
    fecha: str = Form(...),
    inicio: str = Form(""),
    fin: str = Form(""),
    horas: str = Form(""),
    observaciones: str = Form(""),
):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO turnos (trabajador_id, fecha, inicio_jornada, fin_jornada, total_horas, observaciones)
            VALUES (:t, :f, :i, :fi, :h, :o)
        """), {
            "t": int(trabajador_id),
            "f": fecha,
            "i": inicio or None,
            "fi": fin or None,
            "h": horas or None,
            "o": observaciones.strip() or None,
        })

    return RedirectResponse(url="/turnos", status_code=303)


@app.get("/turnos/exportar-excel")
def exportar_turnos_excel(request: Request):
    username, rol, response = require_login(request)
    if response:
        return response

    with engine.begin() as conn:
        turnos = conn.execute(text("""
            SELECT tu.id, tu.fecha, tu.turno, tu.hora_inicio, tu.hora_fin, tu.estado, tu.observaciones,
                   z.nombre AS zona, tr.nombre, tr.apellidos
            FROM turnos tu
            LEFT JOIN trabajadores tr ON tr.id = tu.trabajador_id
            LEFT JOIN zonas z ON z.id = tu.zona_id
            ORDER BY tu.fecha DESC, tu.hora_inicio ASC, tu.id DESC
        """)).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Turnos"

    rojo = PatternFill("solid", fgColor="C00000")
    negro = PatternFill("solid", fgColor="111111")
    gris = PatternFill("solid", fgColor="F2F2F2")
    blanca = Font(color="FFFFFF", bold=True)
    negra = Font(color="000000", bold=True)
    titulo = Font(size=16, bold=True)
    borde = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    centrado = Alignment(horizontal="center", vertical="center")
    izquierda = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("A1:H1")
    ws["A1"] = "LISTADO DE TURNOS"
    ws["A1"].font = titulo
    ws["A1"].alignment = centrado

    ws["A3"] = "Total turnos"
    ws["B3"] = len(turnos)

    ws["A3"].fill = negro
    ws["A3"].font = blanca
    ws["A3"].border = borde
    ws["A3"].alignment = izquierda

    ws["B3"].fill = gris
    ws["B3"].font = negra
    ws["B3"].border = borde
    ws["B3"].alignment = izquierda

    encabezados = ["ID", "Empleado", "Fecha", "Turno", "Inicio", "Fin", "Zona", "Estado"]
    fila = 5
    for i, e in enumerate(encabezados, start=1):
        c = ws.cell(row=fila, column=i, value=e)
        c.fill = rojo
        c.font = blanca
        c.border = borde
        c.alignment = centrado

    fila += 1
    for tu in turnos:
        empleado = f"{tu['nombre'] or ''} {tu['apellidos'] or ''}".strip()
        valores = [
            tu["id"],
            empleado,
            str(tu["fecha"] or ""),
            tu["turno"] or "",
            str(tu["hora_inicio"] or ""),
            str(tu["hora_fin"] or ""),
            tu["zona"] or "",
            tu["estado"] or "",
        ]
        for col, val in enumerate(valores, start=1):
            c = ws.cell(row=fila, column=col, value=val)
            c.border = borde
            c.alignment = centrado if col in (1, 3, 4, 5, 6, 8) else izquierda
        fila += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 24
    ws.column_dimensions["H"].width = 16

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=turnos.xlsx"},
    )


@app.get("/turnos/{id}", response_class=HTMLResponse)
def detalle_turno(request: Request, id: int):
    username, rol, response = require_login(request)
    if response:
        return response

    with engine.begin() as conn:
        turno = conn.execute(text("""
            SELECT tu.*, tr.nombre, tr.apellidos
            FROM turnos tu
            LEFT JOIN trabajadores tr ON tr.id = tu.trabajador_id
            WHERE tu.id = :id
        """), {"id": id}).mappings().first()

        lineas = conn.execute(
            text("SELECT * FROM turno_lineas WHERE turno_id = :id"),
            {"id": id},
        ).mappings().all()

    return render_template(request, "turno_detalle.html", {
        "turno": turno,
        "lineas": lineas,
        "username": username,
        "rol": rol,
        "active_page": "turnos",
    })


@app.post("/turnos/{id}/add_linea")
def add_linea(request: Request, id: int, hora_inicio: str = Form(...), hora_fin: str = Form(...), tarea: str = Form(...)):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        conn.execute(
            text("INSERT INTO turno_lineas (turno_id, hora_inicio, hora_fin, tarea) VALUES (:t, :hi, :hf, :ta)"),
            {"t": id, "hi": hora_inicio, "hf": hora_fin, "ta": tarea.strip()},
        )

    return RedirectResponse(url=f"/turnos/{id}", status_code=303)


@app.get("/turnos/{id}/del_linea/{lid}")
def del_linea(request: Request, id: int, lid: int):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        conn.execute(text("DELETE FROM turno_lineas WHERE id = :id"), {"id": lid})

    return RedirectResponse(url=f"/turnos/{id}", status_code=303)


@app.get("/turnos/delete/{id}")
def eliminar_turno(request: Request, id: int):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        conn.execute(text("DELETE FROM turnos WHERE id = :id"), {"id": id})

    return RedirectResponse(url="/turnos", status_code=303)


@app.get("/rutas", response_class=HTMLResponse)
def ver_rutas(request: Request):
    username, rol, response = require_login(request)
    if response:
        return response

    with engine.begin() as conn:
        rutas = conn.execute(text("""
            SELECT r.id, r.trabajador_id, t.nombre AS trabajador_nombre, t.apellidos AS trabajador_apellidos,
                   r.fecha, r.hora_inicio_jornada, r.hora_fin_jornada, r.total_horas, r.observaciones_generales
            FROM rutas r
            INNER JOIN trabajadores t ON r.trabajador_id = t.id
            ORDER BY r.fecha DESC, r.hora_inicio_jornada ASC
        """)).mappings().all()

        trabajadores = conn.execute(text("""
            SELECT id, nombre, apellidos
            FROM trabajadores
            ORDER BY nombre, apellidos
        """)).mappings().all()

    return render_template(request, "rutas.html", {
        "username": username,
        "rol": rol,
        "rutas": rutas,
        "trabajadores": trabajadores,
        "active_page": "rutas",
    })


@app.get("/rutas/exportar")
def exportar_rutas_excel(request: Request):
    username, rol, response = require_login(request)
    if response:
        return response

    with engine.begin() as conn:
        rutas = conn.execute(text("""
            SELECT
                r.id,
                r.fecha,
                r.hora_inicio_jornada,
                r.hora_fin_jornada,
                r.total_horas,
                r.observaciones_generales,
                t.nombre,
                t.apellidos
            FROM rutas r
            LEFT JOIN trabajadores t ON t.id = r.trabajador_id
            ORDER BY r.fecha DESC, r.hora_inicio_jornada ASC, r.id DESC
        """)).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Rutas"

    rojo = PatternFill("solid", fgColor="C00000")
    negro = PatternFill("solid", fgColor="111111")
    gris = PatternFill("solid", fgColor="F2F2F2")
    blanca = Font(color="FFFFFF", bold=True)
    negra = Font(color="000000", bold=True)
    titulo = Font(size=16, bold=True)
    borde = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    centrado = Alignment(horizontal="center", vertical="center")
    izquierda = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("A1:G1")
    ws["A1"] = "LISTADO DE RUTAS / JORNADAS"
    ws["A1"].font = titulo
    ws["A1"].alignment = centrado

    ws["A3"] = "Total rutas"
    ws["B3"] = len(rutas)

    ws["A3"].fill = negro
    ws["A3"].font = blanca
    ws["A3"].border = borde
    ws["A3"].alignment = izquierda

    ws["B3"].fill = gris
    ws["B3"].font = negra
    ws["B3"].border = borde
    ws["B3"].alignment = izquierda

    encabezados = ["ID", "Empleado", "Fecha", "Inicio", "Fin", "Horas", "Observaciones"]
    fila = 5

    for i, e in enumerate(encabezados, start=1):
        c = ws.cell(row=fila, column=i, value=e)
        c.fill = rojo
        c.font = blanca
        c.border = borde
        c.alignment = centrado

    fila += 1
    for r in rutas:
        empleado = f"{r['nombre'] or ''} {r['apellidos'] or ''}".strip()
        valores = [
            r["id"],
            empleado,
            str(r["fecha"] or ""),
            str(r["hora_inicio_jornada"] or ""),
            str(r["hora_fin_jornada"] or ""),
            r["total_horas"] if r["total_horas"] is not None else "",
            r["observaciones_generales"] or "",
        ]

        for col, val in enumerate(valores, start=1):
            c = ws.cell(row=fila, column=col, value=val)
            c.border = borde
            c.alignment = centrado if col in (1, 3, 4, 5, 6) else izquierda

        fila += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 40

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=rutas.xlsx"},
    )


@app.post("/rutas/nueva")
def crear_ruta_web(
    request: Request,
    trabajador_id: str = Form(...),
    fecha: str = Form(...),
    inicio_jornada: str = Form(...),
    fin_jornada: str = Form(...),
    total_horas: str = Form("8.00"),
    observaciones: str = Form(""),
):
    username, rol, response = require_admin(request)
    if response:
        return response

    try:
        with engine.begin() as conn:
            conn.execute(text("""
                INSERT INTO rutas (
                    trabajador_id,
                    fecha,
                    hora_inicio_jornada,
                    hora_fin_jornada,
                    total_horas,
                    observaciones_generales
                ) VALUES (
                    :trabajador_id,
                    :fecha,
                    :inicio,
                    :fin,
                    :total_horas,
                    :observaciones
                )
            """), {
                "trabajador_id": int(trabajador_id),
                "fecha": fecha,
                "inicio": inicio_jornada,
                "fin": fin_jornada,
                "total_horas": float(total_horas),
                "observaciones": observaciones.strip() or None,
            })
        return RedirectResponse(url="/rutas", status_code=303)
    except Exception as e:
        print(f"[RUTAS NUEVA ERROR] {e}")
        return RedirectResponse(url="/rutas", status_code=303)


@app.get("/rutas/{id}", response_class=HTMLResponse)
def detalle_ruta(request: Request, id: int):
    username, rol, response = require_login(request)
    if response:
        return response

    with engine.begin() as conn:
        ruta = conn.execute(text("""
            SELECT r.id, r.trabajador_id, t.nombre AS trabajador_nombre, t.apellidos AS trabajador_apellidos,
                   r.fecha, r.hora_inicio_jornada, r.hora_fin_jornada, r.total_horas, r.observaciones_generales
            FROM rutas r
            INNER JOIN trabajadores t ON r.trabajador_id = t.id
            WHERE r.id = :id
        """), {"id": id}).mappings().first()

        if not ruta:
            return RedirectResponse(url="/rutas", status_code=303)

        lineas = conn.execute(text("""
            SELECT id, ruta_id, zona, hora_inicio, hora_fin, duracion_minutos, estado, observaciones
            FROM rutas_detalle
            WHERE ruta_id = :id
            ORDER BY hora_inicio ASC
        """), {"id": id}).mappings().all()

    return render_template(request, "ruta_detalle.html", {
        "username": username,
        "rol": rol,
        "ruta": ruta,
        "lineas": lineas,
        "active_page": "rutas",
    })


@app.post("/rutas/{id}/lineas/nueva")
def crear_ruta_linea_web(
    request: Request,
    id: int,
    hora_inicio: str = Form(...),
    hora_fin: str = Form(...),
    tarea: str = Form(...),
):
    username, rol, response = require_admin(request)
    if response:
        return response

    try:
        with engine.begin() as conn:
            conn.execute(text("""
                INSERT INTO rutas_detalle (
                    ruta_id,
                    zona,
                    hora_inicio,
                    hora_fin,
                    duracion_minutos,
                    estado,
                    observaciones
                ) VALUES (
                    :ruta_id,
                    :zona,
                    :hora_inicio,
                    :hora_fin,
                    0,
                    'pendiente',
                    NULL
                )
            """), {
                "ruta_id": id,
                "zona": tarea.strip(),
                "hora_inicio": hora_inicio,
                "hora_fin": hora_fin,
            })
        return RedirectResponse(url=f"/rutas/{id}", status_code=303)
    except Exception as e:
        print(f"[RUTAS LINEA ERROR] {e}")
        return RedirectResponse(url=f"/rutas/{id}", status_code=303)


@app.get("/rutas/{id}/lineas/eliminar/{linea_id}")
def eliminar_ruta_linea_web(request: Request, id: int, linea_id: int):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        conn.execute(text("DELETE FROM rutas_detalle WHERE id = :id"), {"id": linea_id})

    return RedirectResponse(url=f"/rutas/{id}", status_code=303)


@app.get("/rutas/eliminar/{id}")
def eliminar_ruta_web(request: Request, id: int):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        conn.execute(text("DELETE FROM rutas WHERE id = :id"), {"id": id})

    return RedirectResponse(url="/rutas", status_code=303)


@app.get("/horas", response_class=HTMLResponse)
def ver_horas(request: Request):
    username, rol, response = require_login(request)
    if response:
        return response

    with engine.begin() as conn:
        horas = conn.execute(text("""
            SELECT h.id, t.nombre, t.apellidos, h.fecha, h.horas, h.horas_extra, h.tipo
            FROM horas_trabajadas h
            JOIN trabajadores t ON h.trabajador_id = t.id
            ORDER BY h.fecha DESC, h.id DESC
        """)).mappings().all()

        trabajadores = conn.execute(text("""
            SELECT id, nombre, apellidos
            FROM trabajadores
            ORDER BY nombre, apellidos
        """)).mappings().all()

    return render_template(request, "horas.html", {
        "username": username,
        "rol": rol,
        "horas": horas,
        "trabajadores": trabajadores,
        "active_page": "horas",
    })


@app.post("/horas/nueva")
def crear_horas_web(
    request: Request,
    trabajador_id: str = Form(...),
    fecha: str = Form(...),
    horas: str = Form(...),
    horas_extra: str = Form("0"),
    tipo: str = Form("normal"),
):
    username, rol, response = require_admin(request)
    if response:
        return response

    try:
        with engine.begin() as conn:
            conn.execute(text("""
                INSERT INTO horas_trabajadas (trabajador_id, fecha, horas, horas_extra, tipo)
                VALUES (:trabajador_id, :fecha, :horas, :horas_extra, :tipo)
            """), {
                "trabajador_id": int(trabajador_id),
                "fecha": fecha,
                "horas": float(horas),
                "horas_extra": float(horas_extra or 0),
                "tipo": tipo.strip() or "normal",
            })
        return RedirectResponse(url="/horas", status_code=303)
    except Exception as e:
        print(f"[HORAS NUEVA ERROR] {e}")
        return RedirectResponse(url="/horas", status_code=303)


@app.get("/horas/eliminar/{id}")
def eliminar_horas_web(request: Request, id: int):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        conn.execute(text("DELETE FROM horas_trabajadas WHERE id = :id"), {"id": id})

    return RedirectResponse(url="/horas", status_code=303)


@app.get("/ausencias", response_class=HTMLResponse)
def ver_ausencias(request: Request):
    username, rol, response = require_login(request)
    if response:
        return response

    with engine.begin() as conn:
        ausencias = conn.execute(text("""
            SELECT a.id, t.nombre, t.apellidos, a.tipo, a.fecha_inicio, a.fecha_fin, a.motivo, a.estado
            FROM ausencias a
            JOIN trabajadores t ON a.trabajador_id = t.id
            ORDER BY a.fecha_inicio DESC, a.id DESC
        """)).mappings().all()

        trabajadores = conn.execute(text("""
            SELECT id, nombre, apellidos
            FROM trabajadores
            ORDER BY nombre, apellidos
        """)).mappings().all()

    return render_template(request, "ausencias.html", {
        "username": username,
        "rol": rol,
        "ausencias": ausencias,
        "trabajadores": trabajadores,
        "active_page": "ausencias",
    })


@app.post("/ausencias/nueva")
def crear_ausencia_web(
    request: Request,
    trabajador_id: str = Form(...),
    tipo: str = Form(...),
    fecha_inicio: str = Form(...),
    fecha_fin: str = Form(...),
    motivo: str = Form(""),
    estado: str = Form("pendiente"),
):
    username, rol, response = require_admin(request)
    if response:
        return response

    try:
        with engine.begin() as conn:
            conn.execute(text("""
                INSERT INTO ausencias (trabajador_id, tipo, fecha_inicio, fecha_fin, motivo, estado)
                VALUES (:trabajador_id, :tipo, :fecha_inicio, :fecha_fin, :motivo, :estado)
            """), {
                "trabajador_id": int(trabajador_id),
                "tipo": tipo.strip(),
                "fecha_inicio": fecha_inicio,
                "fecha_fin": fecha_fin,
                "motivo": motivo.strip() or None,
                "estado": estado.strip() or "pendiente",
            })
        return RedirectResponse(url="/ausencias", status_code=303)
    except Exception as e:
        print(f"[AUSENCIAS NUEVA ERROR] {e}")
        return RedirectResponse(url="/ausencias", status_code=303)


@app.get("/ausencias/eliminar/{id}")
def eliminar_ausencia_web(request: Request, id: int):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        conn.execute(text("DELETE FROM ausencias WHERE id = :id"), {"id": id})

    return RedirectResponse(url="/ausencias", status_code=303)


@app.get("/productos", response_class=HTMLResponse)
def ver_productos(request: Request):
    username, rol, response = require_login(request)
    if response:
        return response

    with engine.begin() as conn:
        productos = conn.execute(text("""
            SELECT id, nombre, categoria, stock
            FROM productos
            ORDER BY id DESC
        """)).mappings().all()

    return render_template(request, "productos.html", {
        "username": username,
        "rol": rol,
        "productos": productos,
        "active_page": "productos",
    })


@app.post("/productos/nuevo")
def crear_producto_web(
    request: Request,
    nombre: str = Form(...),
    categoria: str = Form(""),
    stock: str = Form("0"),
):
    username, rol, response = require_admin(request)
    if response:
        return response

    try:
        with engine.begin() as conn:
            conn.execute(text("""
                INSERT INTO productos (nombre, categoria, stock)
                VALUES (:nombre, :categoria, :stock)
            """), {
                "nombre": nombre.strip(),
                "categoria": categoria.strip() or None,
                "stock": int(stock or 0),
            })
        return RedirectResponse(url="/productos", status_code=303)
    except Exception as e:
        print(f"[PRODUCTOS NUEVO ERROR] {e}")
        return RedirectResponse(url="/productos", status_code=303)


@app.get("/productos/eliminar/{id}")
def eliminar_producto_web(request: Request, id: int):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        conn.execute(text("DELETE FROM productos WHERE id = :id"), {"id": id})

    return RedirectResponse(url="/productos", status_code=303)


@app.get("/productos/restar/{id}")
def restar_producto(request: Request, id: int):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        conn.execute(
            text("UPDATE productos SET stock = GREATEST(stock - 1, 0) WHERE id = :id"),
            {"id": id},
        )

    return RedirectResponse(url="/productos", status_code=303)


@app.get("/zonas", response_class=HTMLResponse)
def ver_zonas(request: Request):
    username, rol, response = require_login(request)
    if response:
        return response

    with engine.begin() as conn:
        zonas = conn.execute(text("""
            SELECT id, nombre, tipo, activo
            FROM zonas
            ORDER BY nombre ASC, id DESC
        """)).mappings().all()

    return render_template(request, "zonas.html", {
        "username": username,
        "rol": rol,
        "zonas": zonas,
        "active_page": "zonas",
    })


@app.post("/zonas/nueva")
def crear_zona_web(
    request: Request,
    nombre: str = Form(...),
    tipo: str = Form(""),
    activo: str = Form("true"),
):
    username, rol, response = require_admin(request)
    if response:
        return response

    try:
        with engine.begin() as conn:
            conn.execute(text("""
                INSERT INTO zonas (nombre, tipo, activo)
                VALUES (:nombre, :tipo, :activo)
            """), {
                "nombre": nombre.strip(),
                "tipo": tipo.strip() or None,
                "activo": activo.lower() == "true",
            })
        return RedirectResponse(url="/zonas", status_code=303)
    except Exception as e:
        print(f"[ZONAS NUEVA ERROR] {e}")
        return RedirectResponse(url="/zonas", status_code=303)


@app.get("/zonas/toggle/{id}")
def toggle_zona_web(request: Request, id: int):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        conn.execute(text("""
            UPDATE zonas
            SET activo = NOT activo
            WHERE id = :id
        """), {"id": id})

    return RedirectResponse(url="/zonas", status_code=303)


@app.get("/zonas/eliminar/{id}")
def eliminar_zona_web(request: Request, id: int):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        conn.execute(text("DELETE FROM zonas WHERE id = :id"), {"id": id})

    return RedirectResponse(url="/zonas", status_code=303)


@app.get("/productos/exportar")
def exportar_productos_excel(request: Request):
    username, rol, response = require_login(request)
    if response:
        return response

    with engine.begin() as conn:
        productos = conn.execute(text("""
            SELECT id, nombre, categoria, stock
            FROM productos
            ORDER BY id DESC
        """)).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    rojo = PatternFill("solid", fgColor="C00000")
    negro = PatternFill("solid", fgColor="111111")
    gris = PatternFill("solid", fgColor="F2F2F2")
    blanca = Font(color="FFFFFF", bold=True)
    negra = Font(color="000000", bold=True)
    titulo = Font(size=16, bold=True)
    borde = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    centrado = Alignment(horizontal="center", vertical="center")
    izquierda = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A1:D1")
    ws["A1"] = "LISTADO DE PRODUCTOS"
    ws["A1"].font = titulo
    ws["A1"].alignment = centrado

    ws["A3"] = "Total productos"
    ws["B3"] = len(productos)

    ws["A3"].fill = negro
    ws["A3"].font = blanca
    ws["A3"].border = borde
    ws["A3"].alignment = izquierda

    ws["B3"].fill = gris
    ws["B3"].font = negra
    ws["B3"].border = borde
    ws["B3"].alignment = izquierda

    encabezados = ["ID", "Nombre", "Categoría", "Stock"]
    fila = 5
    for i, e in enumerate(encabezados, start=1):
        c = ws.cell(row=fila, column=i, value=e)
        c.fill = rojo
        c.font = blanca
        c.border = borde
        c.alignment = centrado

    fila += 1
    for prod in productos:
        valores = [
            prod["id"],
            prod["nombre"],
            prod["categoria"] or "",
            prod["stock"],
        ]
        for col, val in enumerate(valores, start=1):
            c = ws.cell(row=fila, column=col, value=val)
            c.border = borde
            c.alignment = izquierda if col in (2, 3) else centrado
        fila += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 12

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=productos.xlsx"},
    )


@app.get("/tareas/exportar")
def exportar_tareas_excel(request: Request):
    username, rol, response = require_login(request)
    if response:
        return response

    with engine.begin() as conn:
        tareas = conn.execute(text("""
            SELECT ta.id, ta.titulo, ta.descripcion, ta.estado, ta.prioridad,
                   ta.fecha_asignacion, ta.fecha_vencimiento,
                   tr.nombre, tr.apellidos
            FROM tareas ta
            LEFT JOIN trabajadores tr ON tr.id = ta.trabajador_id
            ORDER BY ta.fecha_asignacion DESC, ta.id DESC
        """)).mappings().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Tareas"

    rojo = PatternFill("solid", fgColor="C00000")
    negro = PatternFill("solid", fgColor="111111")
    gris = PatternFill("solid", fgColor="F2F2F2")
    blanca = Font(color="FFFFFF", bold=True)
    negra = Font(color="000000", bold=True)
    titulo = Font(size=16, bold=True)
    borde = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    centrado = Alignment(horizontal="center", vertical="center")
    izquierda = Alignment(horizontal="left", vertical="center", wrap_text=True)

    ws.merge_cells("A1:H1")
    ws["A1"] = "LISTADO DE TAREAS"
    ws["A1"].font = titulo
    ws["A1"].alignment = centrado

    ws["A3"] = "Total tareas"
    ws["B3"] = len(tareas)

    ws["A3"].fill = negro
    ws["A3"].font = blanca
    ws["A3"].border = borde
    ws["A3"].alignment = izquierda

    ws["B3"].fill = gris
    ws["B3"].font = negra
    ws["B3"].border = borde
    ws["B3"].alignment = izquierda

    encabezados = ["ID", "Título", "Descripción", "Trabajador", "Estado", "Prioridad", "Asignación", "Vencimiento"]
    fila = 5
    for i, e in enumerate(encabezados, start=1):
        c = ws.cell(row=fila, column=i, value=e)
        c.fill = rojo
        c.font = blanca
        c.border = borde
        c.alignment = centrado

    fila += 1
    for t in tareas:
        trabajador = f"{t['nombre'] or ''} {t['apellidos'] or ''}".strip()
        valores = [
            t["id"],
            t["titulo"] or "",
            t["descripcion"] or "",
            trabajador,
            t["estado"] or "",
            t["prioridad"] or "",
            str(t["fecha_asignacion"] or ""),
            str(t["fecha_vencimiento"] or ""),
        ]
        for col, val in enumerate(valores, start=1):
            c = ws.cell(row=fila, column=col, value=val)
            c.border = borde
            c.alignment = centrado if col in (1, 5, 6, 7, 8) else izquierda
        fila += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 16

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=tareas.xlsx"},
    )


@app.get("/departamentos", response_class=HTMLResponse)
def ver_departamentos(request: Request):
    username, rol, response = require_login(request)
    if response:
        return response

    with engine.begin() as conn:
        departamentos = conn.execute(text("""
            SELECT id, nombre
            FROM departamentos
            WHERE activo = true
            ORDER BY nombre ASC
        """)).mappings().all()

    return render_template(request, "departamentos.html", {
        "username": username,
        "rol": rol,
        "departamentos": departamentos,
        "active_page": "departamentos",
    })


@app.get("/ai/logs", response_class=HTMLResponse)
def ai_logs(request: Request):
    username, rol, response = require_admin(request)
    if response:
        return response

    with engine.begin() as conn:
        logs = conn.execute(
            text("SELECT id, usuario, accion, detalle, fecha FROM ai_logs ORDER BY id DESC LIMIT 100")
        ).mappings().all()

    return render_template(request, "ai_logs.html", {
        "logs": logs,
        "username": username,
        "rol": rol,
        "active_page": "ai_logs",
    })